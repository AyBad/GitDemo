import openpyxl
Dict = {}
book = openpyxl.load_workbook("C:\\Users\\p104981\\PycharmProjects\\pythonSelfFramework\\TestDatas\\Datas.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
#print(cell.value)

sheet.cell(row=2, column=2).value = "ayman"
#print (sheet.cell(row=2, column=2).value)
#print(sheet.max_row)
#print (sheet.max_column)

#print (sheet['A2'].value)

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "TestCase1":
        for j in range(2, sheet.max_column+1):
            #print (sheet.cell(row=i, column=j).value)
            # Dict["email"] = ayman.badri.tn@gmail.com
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print (Dict)
print ([Dict])


