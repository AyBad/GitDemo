import openpyxl


class HomePageData:

    test_home_page_data = [{"firstname":"Ayman", "email":"badri.ayman.tn@gmail.com", "gender":"Male"},
                            {"firstname":"yasmine", "email":"yasminebenamor97", "gender":"Female"}]
    @staticmethod
    def GetTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\p104981\\PycharmProjects\\pythonSelfFramework\\TestDatas\\Datas.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i,column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    # print (sheet.cell(row=i, column=j).value)
                    # Dict["email"] = ayman.badri.tn@gmail.com
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]